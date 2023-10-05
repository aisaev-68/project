import io
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment
import pandas as pd
from typing import List


async def process_excel_data(file: bytes) -> List:
    """
    Функция для загрузки файла Excel в DataFrame с помощью Pandas
    :param filename: название файла.
    :return: список данных.
    """
    bytes_io = io.BytesIO(file)
    df = pd.read_excel(bytes_io, sheet_name='data', header=[0, 1])

    data = []

    for _, row in df.iterrows():
        project_id, project_name = int(row[('Unnamed: 0_level_0', 'Код')]), row[
            ('Unnamed: 1_level_0', 'Наименование проекта')]
        project_data = {'code': project_id, 'project': project_name, 'values': []}

        # Перебираем столбцы с данными (начиная с 3-го)
        for i in range(2, len(row), 2):
            date = df.columns[i][0]
            plan = row[df.columns[i][0], 'план']
            fact = row[df.columns[i][0], 'факт']

            plan = None if pd.isna(plan) else plan
            fact = None if pd.isna(fact) else fact

            project_data['values'].append({'date': date, 'plan_value': plan, 'fact_value': fact})

        data.append(project_data)

    return data


def write_to_excel(db_data: list):
    """
    Функция для записи файла и передачи его для выгрузки.
    :param data: данные из базы.
    :param version_id: версия файла.
    :return: данные в байтах для скачивания.
    """

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 25

    sheet.append(['', '', '', ''])

    first_data_item = db_data[0]

    dates = [value_item['date'] for value_item in first_data_item['values']]
    for i in range(len(dates)):
        sheet.cell(row=1, column=i * 2 + 3).value = dates[i]
        sheet.merge_cells(start_row=1, start_column=i * 2 + 3, end_row=1, end_column=i * 2 + 4)

    header_row = ['Код', 'Наименование проекта'] + ['план', 'факт'] * len(dates)
    sheet.append(header_row)

    for item in db_data:
        code = item['code']
        project = item['project']
        plan_values = [value_item['plan_value'] for value_item in item['values']]
        fact_values = [value_item['fact_value'] for value_item in item['values']]
        sheet.append([code, project] + plan_values + fact_values)

    for row in sheet.iter_rows(min_row=1, max_row=len(db_data) + 3, min_col=1, max_col=len(dates) * 2 + 2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    sheet.title = 'data'
    excel_output = BytesIO()
    workbook.save(excel_output)

    excel_output.seek(0)
    excel_bytes = excel_output.read()

    return excel_bytes
